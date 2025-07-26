from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions
from django.db.models import Count, Q
from django.db.models.functions import TruncMonth
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from scheduling.models import Aula
from scheduling.filters import AulaFilter
from users.models import CustomUser
from .serializers import AdminDashboardSerializer


class AdminDashboardAPIView(APIView):
    """
    Endpoint de leitura que agrega dados de todo o sistema para
    um dashboard de administrador.
    """
    permission_classes = [permissions.IsAdminUser]

    def get(self, request, *args, **kwargs):
        params = request.query_params
        aulas_queryset = Aula.objects.all()

        if params.get('data_inicial'):
            aulas_queryset = aulas_queryset.filter(data_hora__date__gte=params.get('data_inicial'))
        if params.get('data_final'):
            aulas_queryset = aulas_queryset.filter(data_hora__date__lte=params.get('data_final'))

        total_aulas = aulas_queryset.count()
        total_realizadas = aulas_queryset.filter(status="Realizada").count()
        total_canceladas = aulas_queryset.filter(status="Cancelada").count()
        total_aluno_ausente = aulas_queryset.filter(status="Aluno Ausente").count()
        aulas_concluidas = total_realizadas + total_aluno_ausente
        taxa_sucesso = (total_realizadas / aulas_concluidas * 100) if aulas_concluidas > 0 else 0

        cat_chart_data = aulas_queryset.values('modalidade__nome').annotate(contagem=Count('id')).order_by('-contagem')
        mes_chart_data = aulas_queryset.filter(status='Realizada').annotate(mes=TruncMonth('data_hora')).values('mes').annotate(contagem=Count('id')).order_by('mes')

        professores_qs = CustomUser.objects.filter(tipo__in=['professor', 'admin']).annotate(
            total_atribuidas=Count('aulas', filter=Q(aulas__in=aulas_queryset), distinct=True),
            total_realizadas=Count('relatorios_validados', filter=Q(relatorios_validados__aula__in=aulas_queryset), distinct=True)
        ).filter(total_atribuidas__gt=0).order_by('-total_realizadas')

        prof_performance_data = []
        for p in professores_qs:
            taxa = (p.total_realizadas / p.total_atribuidas * 100) if p.total_atribuidas > 0 else 0
            prof_performance_data.append({
                'username': p.username,
                'total_realizadas': p.total_realizadas,
                'total_atribuidas': p.total_atribuidas,
                'taxa_realizacao_percentual': round(taxa, 2)
            })

        data = {
            'kpis': {
                'total_aulas': total_aulas,
                'total_realizadas': total_realizadas,
                'total_canceladas': total_canceladas,
                'total_aluno_ausente': total_aluno_ausente,
                'taxa_sucesso_percentual': round(taxa_sucesso, 2)
            },
            'aulas_por_categoria_chart': {
                'labels': [item['modalidade__nome'] for item in cat_chart_data],
                'data': [item['contagem'] for item in cat_chart_data]
            },
            'aulas_realizadas_por_mes_chart': {
                'labels': [item['mes'].strftime('%b/%Y') for item in mes_chart_data],
                'data': [item['contagem'] for item in mes_chart_data]
            },
            'professor_performance': prof_performance_data
        }

        serializer = AdminDashboardSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        return Response(serializer.validated_data)


class ExportAulasAPIView(APIView):
    """
    Endpoint para exportar uma lista filtrada de aulas para um arquivo Excel (.xlsx).
    """
    permission_classes = [permissions.IsAdminUser]

    def get(self, request, *args, **kwargs):
        aulas_queryset = AulaFilter(request.query_params, queryset=Aula.objects.all()).qs
        aulas_list = aulas_queryset.order_by("data_hora").select_related(
            'modalidade', 'relatorio__professor_que_validou'
        ).prefetch_related('alunos', 'professores')

        workbook = Workbook()
        ws = workbook.active
        ws.title = "Relatorio de Aulas"

        headers = [
            "ID Aula", "Data e Hora", "Status", "Modalidade", "Alunos",
            "Prof. Atribuído(s)", "Prof. que Realizou"
        ]
        ws.append(headers)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        for aula in aulas_list:
            alunos_str = ", ".join([al.nome_completo for al in aula.alunos.all()])
            professores_str = ", ".join([p.username for p in aula.professores.all()])
            relatorio = getattr(aula, "relatorio", None)
            professor_realizou_str = relatorio.professor_que_validou.username if relatorio and relatorio.professor_que_validou else "N/A"

            ws.append([
                aula.id,
                aula.data_hora,
                aula.get_status_display(),
                aula.modalidade.nome,
                alunos_str,
                professores_str,
                professor_realizou_str,
            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="relatorio_de_aulas.xlsx"'
        workbook.save(response)

        return response
